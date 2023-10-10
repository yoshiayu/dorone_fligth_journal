from django.shortcuts import render, redirect
from .forms import FlightRecordForm
from django.http import JsonResponse
from .models import FlightRecord
from django.views.decorators.csrf import csrf_exempt
import json
from django.contrib.auth.decorators import login_required
import openpyxl
from django.http import HttpResponse

def export_to_excel(request):
    # ファイルを開く
    wb = openpyxl.load_workbook('static/飛行日誌.xlsx')
    ws = wb.active

    # データをセルに書き込む
    ws['B6'] = request.POST['flight_date']
    ws['C6'] = request.POST['flyer_name']
    # ... 他のデータも同様に書き込む ...

    # ファイル名を設定してレスポンスを返す
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=new_飛行日誌.xlsx'
    wb.save(response)

    return JsonResponse({'message': 'Success'})
    # return response

@csrf_exempt
def save_record(request):
    if request.method == "POST":
        data = json.loads(request.body)
        record = FlightRecord(**data)
        record.save()
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "fail"})

@login_required
def flight_record(request):
    if request.method == "POST":
        form = FlightRecordForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, 'index.html', {})
    else:
        form = FlightRecordForm()
    return render(request, 'flight_record/index.html', {'form': form})
